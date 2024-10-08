import os
from datetime import datetime, timedelta

import requests
import json
import pandas as pd

import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment


# Функция для получения списка рекламных кампаний со статусом 9
def get_active_advertisements(api_key):
    url = "https://advert-api.wildberries.ru/adv/v1/promotion/count"
    headers = {
        "Authorization": api_key
    }
    response = requests.get(url, headers=headers)
    print(response.status_code)
    adverts = response.json().get('adverts', [])

    advert_ids = []
    for advert in adverts:
        if advert.get('status') == 9:
            advert_list = advert.get('advert_list', [])
            for item in advert_list:
                advert_ids.append(item.get('advertId'))

    return advert_ids


def get_article_map(api_key):
    url = "http://217.25.93.96/root_id?company=missyourkiss"
    headers = {
        "api-key": api_key
    }
    response = requests.get(url, headers=headers)

    article_map = {}

    if response.status_code == 200:
        try:
            data = response.json()
            for item in data:
                article_map[str(item['nm'])] = item['article']
        except json.JSONDecodeError as e:
            print(f"JSONDecodeError: {e}")
    else:
        print(f"Error fetching articles: {response.status_code}")

    return article_map


def get_advert_statistics(company_api_key, api_key, authorizev3, cookie, user_agent, advert_id, start_date, end_date):
    # Получение маппинга nm_id -> article
    article_map = get_article_map(company_api_key)
    print(start_date, end_date)
    url = f"https://cmp.wildberries.ru/api/v5/fullstat?advertID={advert_id}&from={start_date}&to={end_date}&appType=0&placementType=0"
    headers = {
        "Accept-Encoding": "gzip, deflate, br",
        "Authorizev3": authorizev3,
        "Cookie": cookie,
        "User-Agent": user_agent
    }
    response = requests.get(url, headers=headers)

    try:
        print(response.status_code)
        print(response.text)
        data = response.json().get('content', {})
    except json.JSONDecodeError as e:
        print(f"JSONDecodeError for advert_id {advert_id}: {e}")
        return []

    main_items = []
    associated_items = []

    try:
        nm_stats = data.get('nmStats', [])
        side_nm_stats = data.get('sideNmStats', [])

        main_nm_ids = [str(item.get('nm_id')) for item in nm_stats]
        main_nm_names = [str(item.get('name')) for item in nm_stats]
        main_orders = [str(item.get('orders')) for item in nm_stats]

        # Найти артикулы для основного товара
        main_articles = [article_map.get(str(item.get('nm_id')), '0') for item in nm_stats]

        campaign_name = get_campaign_name(authorizev3, cookie, user_agent, advert_id)

        main_item = {
            'advert_id': advert_id,
            'advert_name': campaign_name,
            'main_nm_id': "; ".join(main_nm_ids),
            'main_nm_name': "; ".join(main_nm_names),
            'main_shks': "; ".join(main_orders),
            'main_orders': "; ".join(main_orders),
            'main_article': "; ".join(main_articles)  # Заполняем новую колонку "Артикул товара"
        }
        main_items.append(main_item)

        for item in side_nm_stats:
            associated_article = article_map.get(str(item.get('nm_id')), '0')
            associated_items.append({
                'advert_id': advert_id,
                'advert_name': main_item['advert_name'],
                'main_nm_id': main_item['main_nm_id'],
                'main_nm_name': main_item['main_nm_name'],
                'main_shks': main_item['main_shks'],
                'main_orders': main_item['main_orders'],
                'main_article': main_item['main_article'],
                'associated_nm_id': item.get('nm_id'),
                'associated_nm_name': item.get('name'),
                'associated_orders': item.get('orders'),
                'associated_article': associated_article
                # Заполняем новую колонку "Артикул товара" для ассоциированного товара
            })
    except TypeError as e:
        error_message = f"Error processing advert_id {advert_id}: {e}"
        print(error_message)
    print(main_items)

    return main_items + associated_items


# Основная функция для сбора данных и создания итогового файла Excel
def collect_data(company_api_key, api_key, start_date, end_date, authorizev3, cookie, user_agent):
    advert_ids = get_active_advertisements(api_key)
    print(advert_ids)
    all_stats = []

    for advert_id in advert_ids:
        advert_stats = get_advert_statistics(company_api_key, api_key, authorizev3, cookie, user_agent, advert_id, start_date, end_date)
        if advert_stats:
            all_stats.extend(advert_stats)

        # Секундная задержка между запросами
        time.sleep(10)

    all_stats_df = pd.DataFrame(all_stats)

    # Переименование столбцов
    all_stats_df = all_stats_df.rename(columns={
        'advert_id': 'НОМЕР РК',
        'advert_name': 'НАЗВАНИЕ РК',
        'main_nm_id': 'ОСНОВНОЙ ТОВАР',
        'main_nm_name': 'НАЗВАНИЕ ТОВАРА',
        'main_article': 'АРТИКУЛ ТОВАРА',
        'main_shks': 'ЗАКАЗЫ ОСНОВНОГО ТОВАРА',
        'associated_nm_id': 'АССОЦИИРОВАННЫЙ ТОВАР',
        'associated_nm_name': 'НАЗВАНИЕ АССОЦ.ТОВАРА',
        'associated_article': 'АРТИКУЛ АССОЦ. ТОВАРА',
        'associated_orders': 'ПРОДАЖ АССОЦИИРОВАННОГО ТОВАРА'
    })

    all_stats_df = all_stats_df.drop(columns=['main_orders'])

    # Создание итогового файла Excel
    output_file_path = 'advert_statistics.xlsx'
    all_stats_df.to_excel(output_file_path, index=False)

    # Открытие созданного файла Excel
    workbook = load_workbook(output_file_path)
    worksheet = workbook.active

    # Определяем диапазон для объединения ячеек
    columns_to_merge = ['A', 'B', 'C', 'D', 'E', "F"]
    for col in columns_to_merge:
        cell_value = None
        merge_start = None
        for row in range(2, worksheet.max_row + 1):
            current_value = worksheet[f"{col}{row}"].value
            print(f"Processing column {col}, row {row}, value {current_value}")  # Добавляем отладочный вывод
            if current_value != cell_value:
                if merge_start is not None:
                    print(f"Merging {col}{merge_start}:{col}{row - 1}")  # Добавляем отладочный вывод
                    worksheet.merge_cells(f"{col}{merge_start}:{col}{row - 1}")
                    worksheet[f"{col}{merge_start}"].alignment = Alignment(horizontal="center", vertical="center")
                cell_value = current_value
                merge_start = row
        if merge_start is not None and merge_start != row:
            print(f"Merging {col}{merge_start}:{col}{row}")  # Добавляем отладочный вывод
            worksheet.merge_cells(f"{col}{merge_start}:{col}{row}")
            worksheet[f"{col}{merge_start}"].alignment = Alignment(horizontal="center", vertical="center")

    # Сохранение изменений в файл
    workbook.save(output_file_path)
    return output_file_path


def get_campaign_name(authorizev3, cookie, user_agent, advert_id):
    url = f"https://cmp.wildberries.ru/api/v1/atrevd?advert-id={advert_id}"
    headers = {
        "Accept-Encoding": "gzip, deflate, br",
        "Authorizev3": authorizev3,
        "Cookie": cookie,
        "User-Agent": user_agent
    }

    response = requests.get(url, headers=headers)

    try:
        data = response.json()
        campaign_name = data.get("campaignName", "N/A")
        return campaign_name
    except json.JSONDecodeError as e:
        print(f"JSONDecodeError for advert_id {advert_id}: {e}")
        return "N/A"


# start_date = (datetime.now() - timedelta(days=10)).strftime('%Y-%m-%dT%H:%M:%S') + 'Z'
# end_date = datetime.now().strftime('%Y-%m-%dT%H:%M:%S') + 'Z'
#
# # Замена символов ":" на "%3A"
# start_date = start_date.replace(':', '%3A')
# end_date = end_date.replace(':', '%3A')
#
# api_key = os.getenv('API_TOKEN')
# authorizev3 = os.getenv('authorizev3')
# cookie = os.getenv('COOKIE')
# user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, Gecko) Chrome/124.0.0.0 YaBrowser/24.6.0.0 Safari/537.36"
# company_api_key = os.getenv("MYK_API_KEY")
# collect_data(company_api_key, api_key, start_date, end_date, authorizev3, cookie, user_agent)