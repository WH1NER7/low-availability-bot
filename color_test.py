import openpyxl

COLOR_ORDER = ['#FF0000', '#FF9900', '#FFFF00', '#00FF00', '#00FFFF', '#4A86E8', '#0000FF', '#9900FF', '#FF00FF', '#660000', '#7F6000', '#0C343D', '#20124D', '#E06666', '#FFD966', '#93C47D', '#8E7CC3', '#C27BA0', '#F6B26B', '#A2C4C9']


def get_colors_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    colors = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:  # Проверка на наличие данных в ячейке
                fill = cell.fill
                if fill.start_color.rgb:
                    color = fill.start_color.rgb[-6:]  # Получаем последние 6 символов для hex-кода цвета
                    color = "#" + color  # Преобразуем в формат hex
                    if color in COLOR_ORDER:
                        colors[cell.coordinate] = color
    unique_colors = list(set(colors.values()))
    print("Extracted unique colors:", unique_colors)
    return colors, unique_colors

def add_columns_and_fill_colors(file_path, barcodes, unique_colors):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    num_columns = ws.max_column

    # Переименовываем первые две колонки
    ws.cell(row=1, column=1, value="баркод товара")
    ws.cell(row=1, column=2, value="кол-во товаров")

    # Вставляем 2 новые колонки после существующих
    ws.insert_cols(num_columns + 1)
    ws.insert_cols(num_columns + 2)
    ws.cell(row=1, column=num_columns + 1, value="шк короба")
    ws.cell(row=1, column=num_columns + 2, value="срок годности")

    # Сортируем unique_colors в соответствии с COLOR_ORDER
    color_order_clean = [color[1:] for color in COLOR_ORDER]
    color_index = {color: index for index, color in enumerate(color_order_clean)}
    sorted_colors = sorted([color[1:] for color in unique_colors if color[1:] in color_index], key=lambda color: color_index[color])

    # Соотносим цвета и баркоды
    color_to_barcode = {color: barcodes[i] for i, color in enumerate(sorted_colors) if i < len(barcodes)}
    print("Color to barcode mapping:", color_to_barcode)

    # Заполняем 3ю колонку в зависимости от цвета во второй колонке
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:  # Только ячейки с данными
                fill = cell.fill
                if fill and fill.start_color and fill.start_color.rgb:
                    color = fill.start_color.rgb[-6:]  # Получаем последние 6 символов для hex-кода цвета
                    color = "#" + color  # Преобразуем в формат hex
                    if color not in COLOR_ORDER:
                        print(f"Unexpected color found: {color} in cell {cell.coordinate}")
                    else:
                        barcode = color_to_barcode.get(color[1:], "")
                        ws.cell(row=cell.row, column=num_columns + 1, value=barcode)

    # Сохраняем изменения и возвращаем путь к обновленному файлу
    updated_file_path = "updated_" + file_path
    wb.save(updated_file_path)
    return updated_file_path

# Пример использования
file_path = "book_excel/test1.xlsx"
colors, unique_colors = get_colors_from_excel(file_path)
barcodes = ["barcode1", "barcode2", "barcode3", "barcode4", "barcode5"]  # пример баркодов

updated_file_path = add_columns_and_fill_colors(file_path, barcodes, unique_colors)
print(f"Updated file saved to: {updated_file_path}")
