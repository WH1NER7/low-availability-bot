import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ReportAggregator:
    def __init__(self, file_path, myk_key, delta_threshold=20):
        self.file_path = file_path
        self.myk_key = myk_key
        self.delta_threshold = delta_threshold

    def run(self):
        # Загрузка Excel файла
        df = pd.read_excel(self.file_path)

        # Преобразование столбца с датами в datetime
        df['dt'] = pd.to_datetime(df['dt'])

        # Получение уникальных дат и проверка их кратности двум
        unique_dates = df['dt'].unique()
        if len(unique_dates) % 2 != 0:
            raise ValueError("Количество уникальных дат должно быть кратно двум")

        # Разделение дат на два периода
        first_period_dates = unique_dates[len(unique_dates) // 2:]
        second_period_dates = unique_dates[:len(unique_dates) // 2]

        first_period_str = f"{first_period_dates.min().strftime('%d.%m')}-{first_period_dates.max().strftime('%d.%m')}"
        second_period_str = f"{second_period_dates.min().strftime('%d.%m')}-{second_period_dates.max().strftime('%d.%m')}"

        # Функция для агрегации данных по периодам
        def aggregate_data(df, period_dates):
            period_df = df[df['dt'].isin(period_dates)]
            aggregated = period_df.groupby('nmID').agg(
                ordersCount=('ordersCount', 'sum'),
                buyoutsCount=('buyoutsCount', 'sum')
            ).reset_index()
            return aggregated

        # Агрегация данных по двум периодам
        first_period_aggregated = aggregate_data(df, first_period_dates)
        second_period_aggregated = aggregate_data(df, second_period_dates)

        # Переименуем столбцы для ясности
        first_period_aggregated.columns = ['nmID', 'ordersCount_current', 'buyoutsCount_current']
        second_period_aggregated.columns = ['nmID', 'ordersCount_previous', 'buyoutsCount_previous']

        # Объединение данных по двум периодам
        result_df = pd.merge(first_period_aggregated, second_period_aggregated, on='nmID', how='outer')

        # Преобразование столбцов в числовой формат
        result_df['ordersCount_current'] = pd.to_numeric(result_df['ordersCount_current'], errors='coerce').fillna(0)
        result_df['ordersCount_previous'] = pd.to_numeric(result_df['ordersCount_previous'], errors='coerce').fillna(0)
        result_df['buyoutsCount_current'] = pd.to_numeric(result_df['buyoutsCount_current'], errors='coerce').fillna(0)
        result_df['buyoutsCount_previous'] = pd.to_numeric(result_df['buyoutsCount_previous'], errors='coerce').fillna(0)

        # Фильтрация заказов, чтобы учитывались только те, у которых количество заказов как в текущем, так и в предыдущем периоде больше или равно 10
        result_df = result_df[(result_df['ordersCount_current'] >= 10) & (result_df['ordersCount_previous'] >= 10)]

        # Проверка наличия каждого артикула в каждом из 10 дней
        missing_nmid = []
        for nmID in result_df['nmID']:
            first_period_check = df[(df['nmID'] == nmID) & (df['dt'].isin(first_period_dates))]
            second_period_check = df[(df['nmID'] == nmID) & (df['dt'].isin(second_period_dates))]
            if len(first_period_check) != len(first_period_dates) or len(second_period_check) != len(
                    second_period_dates):
                missing_nmid.append(nmID)

        # Удаление артикулов, которые отсутствуют в некоторых днях
        result_df = result_df[~result_df['nmID'].isin(missing_nmid)]

        # Добавление дельты и процента отклонения
        result_df['orders_delta'] = result_df['ordersCount_current'] - result_df['ordersCount_previous']
        result_df['orders_percent_change'] = (
                (result_df['orders_delta'] / result_df['ordersCount_previous'].replace({0:1})) * 100).round(2)
        result_df['buyouts_delta'] = result_df['buyoutsCount_current'] - result_df['buyoutsCount_previous']
        result_df['buyouts_percent_change'] = (
                (result_df['buyouts_delta'] / result_df['buyoutsCount_previous'].replace({0:1})) * 100).round(2)

        # Фильтрация по проценту изменения
        result_df = result_df[result_df['orders_percent_change'].abs() > self.delta_threshold]

        # Получение данных с API
        api_url = 'http://217.25.93.96/root_id?company=missyourkiss'
        headers = {
            'api-key': self.myk_key
        }
        response = requests.get(api_url, headers=headers)
        items_data = response.json()

        # Преобразование данных API в DataFrame
        items_df = pd.DataFrame(items_data)
        items_df = items_df[['nm', 'item_category', 'item_name', 'article', 'size']]

        items_df = items_df.drop_duplicates(subset='nm')
        items_df.columns = ['nmID', 'Категория', 'Название', 'Артикул', 'Размер']

        # Объединение данных с результатами агрегации
        final_df = pd.merge(items_df, result_df, on='nmID', how='right')

        # Добавление столбца со ссылкой
        final_df['Ссылка'] = 'https://www.wildberries.ru/catalog/' + final_df['nmID'].astype(
            str) + '/detail.aspx?targetUrl=SP'

        # Переупорядочение столбцов
        final_df = final_df[['Категория', 'Название', 'Артикул', 'Размер', 'nmID', 'Ссылка',
                             'ordersCount_current', 'ordersCount_previous', 'orders_delta', 'orders_percent_change',
                             'buyoutsCount_current', 'buyoutsCount_previous', 'buyouts_delta',
                             'buyouts_percent_change']]

        # Переименование столбцов
        final_df.columns = ['Категория', 'Название', 'Артикул', 'Размер', 'nmID', 'Ссылка',
                            f'Заказы текущий период ({first_period_str})',
                            f'Заказы предыдущий период ({second_period_str})',
                            'Дельта заказов', 'Процент отклонения заказов',
                            f'Выкупы текущий период ({first_period_str})',
                            f'Выкупы предыдущий период ({second_period_str})',
                            'Дельта выкупов', 'Процент отклонения выкупов']

        # Сохранение итогового результата в Excel
        output_file_path = f"{first_period_str} - Дельта.xlsx"
        final_df.to_excel(output_file_path, index=False)

        # Открытие созданного Excel файла для добавления стилей
        wb = load_workbook(output_file_path)
        ws = wb.active

        # Определение цветов для заливки
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # Применение заливки для процентного отклонения заказов
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=10, max_row=ws.max_row):
            for cell in row:
                if cell.value > 0:
                    cell.fill = green_fill
                elif cell.value < 0:
                    cell.fill = red_fill

        # Применение заливки для процентного отклонения выкупов
        for row in ws.iter_rows(min_row=2, min_col=14, max_col=14, max_row=ws.max_row):
            for cell in row:
                if cell.value > 0:
                    cell.fill = green_fill
                elif cell.value < 0:
                    cell.fill = red_fill

        # Сохранение файла
        wb.save(output_file_path)

        # Сохранение отсутствующих артикулов в отдельный файл
        missing_nmid_file_path = f"{first_period_str} - missing_nmid.txt"
        with open(missing_nmid_file_path, 'w') as file:
            for nmid in missing_nmid:
                file.write(f"{nmid}\n")

        print(f"Итоговый файл сохранен как {output_file_path}")
        print(f"Список отсутствующих артикулов сохранен как {missing_nmid_file_path}")

        return output_file_path, missing_nmid_file_path
