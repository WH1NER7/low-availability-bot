import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import aiohttp
import asyncio

class OzonReportAggregator:
    def __init__(self, client_id, api_key, delta_threshold=0):
        self.client_id = client_id
        self.api_key = api_key
        self.delta_threshold = delta_threshold
        self.base_url = "https://api-seller.ozon.ru/v1/analytics/data"

    async def get_additional_data(self, url):
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as resp:
                return await resp.json()

    def get_report_data(self, date_from, date_to):
        print(date_from, date_to)
        url = self.base_url

        headers = {
            "Client-Id": self.client_id,
            "Api-Key": self.api_key
        }

        payload = {
            "date_from": date_from,
            "date_to": date_to,
            "metrics": [
                "ordered_units",
                "delivered_units"
            ],
            "dimension": [
                "sku",
                "day"
            ],
            "filters": [],
            "sort": [
                {
                    "key": "ordered_units",
                    "order": "DESC"
                }
            ],
            "limit": 1000,
            "offset": 0
        }

        all_data = []
        offset = 0

        while True:
            payload["offset"] = offset
            response = requests.post(url, json=payload, headers=headers)
            response.raise_for_status()
            data = response.json()

            if 'result' in data and 'data' in data['result']:
                all_data.extend(data['result']['data'])
                if len(data['result']['data']) < 1000:
                    break
                offset += 1000
            else:
                break

        return all_data

    async def run(self, date_start, date_end):
        date_start_strp = datetime.strptime(date_start, '%d.%m.%Y')
        date_end_strp = datetime.strptime(date_end, '%d.%m.%Y')

        date_start_formatted = date_start_strp.strftime('%Y-%m-%d')
        date_end_formatted = date_end_strp.strftime('%Y-%m-%d')

        # Определение предыдущего периода
        period_length = (date_end_strp - date_start_strp).days + 1
        previous_period_start = date_start_strp - timedelta(days=period_length)
        previous_period_end = date_start_strp - timedelta(days=1)

        previous_period_start_formatted = previous_period_start.strftime('%Y-%m-%d')
        previous_period_end_formatted = previous_period_end.strftime('%Y-%m-%d')

        # Получение данных для обоих периодов
        data_current = self.get_report_data(date_start_formatted, date_end_formatted)
        data_previous = self.get_report_data(previous_period_start_formatted, previous_period_end_formatted)

        # Объединение данных обоих периодов
        data = data_current + data_previous

        if not data:
            raise ValueError("Нет данных за указанный период")

        # Преобразование данных в DataFrame
        processed_data = []
        for item in data:
            metrics = item.get('metrics', [0, 0])
            dimensions = item.get('dimensions', [])
            processed_data.append({
                'SKU': str(dimensions[0]['id']) if dimensions else None,  # Преобразование SKU в строку
                'Date': dimensions[1]['id'] if len(dimensions) > 1 else None,
                'Ordered Units': metrics[0],
                'Delivered Units': metrics[1]
            })

        df = pd.DataFrame(processed_data)

        # Проверка наличия необходимых столбцов
        required_columns = {'Date', 'SKU', 'Ordered Units', 'Delivered Units'}
        if not required_columns.issubset(df.columns):
            raise ValueError(f"Отсутствуют необходимые столбцы: {required_columns - set(df.columns)}")

        # Преобразование столбца с датами в datetime
        df['Date'] = pd.to_datetime(df['Date'])

        # Фильтрация данных по текущему и предыдущему периодам
        current_period_dates = pd.date_range(start=date_start_strp, end=date_end_strp)
        previous_period_dates = pd.date_range(start=previous_period_start, end=previous_period_end)

        df_current = df[df['Date'].isin(current_period_dates)]
        df_previous = df[df['Date'].isin(previous_period_dates)]

        current_period_str = f"{date_start_strp.strftime('%d.%m')}-{date_end_strp.strftime('%d.%m')}"
        previous_period_str = f"{previous_period_start.strftime('%d.%m')}-{previous_period_end.strftime('%d.%m')}"

        print(f"Текущий период: {current_period_str}, Предыдущий период: {previous_period_str}")

        # Функция для агрегации данных по периодам
        def aggregate_data(df, period_dates):
            period_df = df[df['Date'].isin(period_dates)]
            aggregated = period_df.groupby('SKU').agg(
                ordered_units=('Ordered Units', 'sum'),
                delivered_units=('Delivered Units', 'sum')
            ).reset_index()
            return aggregated

        # Агрегация данных по двум периодам
        current_period_aggregated = aggregate_data(df, current_period_dates)
        previous_period_aggregated = aggregate_data(df, previous_period_dates)

        print("Агрегированные данные текущего периода:", current_period_aggregated.head())
        print("Агрегированные данные предыдущего периода:", previous_period_aggregated.head())

        # Проверка итоговых сумм по заказам и выкупам перед объединением
        current_period_total_orders = current_period_aggregated['ordered_units'].sum()
        previous_period_total_orders = previous_period_aggregated['ordered_units'].sum()
        current_period_total_deliveries = current_period_aggregated['delivered_units'].sum()
        previous_period_total_deliveries = previous_period_aggregated['delivered_units'].sum()

        print(
            f"Текущий период ({current_period_str}) - Заказы: {current_period_total_orders}, Выкупы: {current_period_total_deliveries}")
        print(
            f"Предыдущий период ({previous_period_str}) - Заказы: {previous_period_total_orders}, Выкупы: {previous_period_total_deliveries}")

        # Продолжение обработки данных...

        # Переименуем столбцы для ясности
        current_period_aggregated.columns = ['SKU', 'ordered_units_current', 'delivered_units_current']
        previous_period_aggregated.columns = ['SKU', 'ordered_units_previous', 'delivered_units_previous']

        # Объединение данных по двум периодам
        result_df = pd.merge(current_period_aggregated, previous_period_aggregated, on='SKU', how='outer')
        print("Объединенные данные:", result_df.head())

        # Преобразование столбцов в числовой формат
        result_df['ordered_units_current'] = pd.to_numeric(result_df['ordered_units_current'], errors='coerce').fillna(0)
        result_df['ordered_units_previous'] = pd.to_numeric(result_df['ordered_units_previous'], errors='coerce').fillna(0)
        result_df['delivered_units_current'] = pd.to_numeric(result_df['delivered_units_current'], errors='coerce').fillna(0)
        result_df['delivered_units_previous'] = pd.to_numeric(result_df['delivered_units_previous'], errors='coerce').fillna(0)

        # Проверка итоговых сумм по заказам и выкупам после объединения
        merged_total_orders = result_df['ordered_units_current'].sum() + result_df['ordered_units_previous'].sum()
        merged_total_deliveries = result_df['delivered_units_current'].sum() + result_df['delivered_units_previous'].sum()

        print(f"Итоговая сумма заказов после объединения: {merged_total_orders}")
        print(f"Итоговая сумма выкупов после объединения: {merged_total_deliveries}")

        # Фильтрация строк с минимальным числом заказов меньше 5
        result_df = result_df[(result_df['ordered_units_current'] >= 5) & (result_df['ordered_units_previous'] >= 5)]

        # Добавление дельты и процента отклонения
        result_df['ordered_units_delta'] = result_df['ordered_units_current'] - result_df['ordered_units_previous']
        result_df['ordered_units_percent_change'] = (
                (result_df['ordered_units_delta'] / result_df['ordered_units_previous'].replace({0: 1})) * 100).round(2)
        result_df['delivered_units_delta'] = result_df['delivered_units_current'] - result_df['delivered_units_previous']
        result_df['delivered_units_percent_change'] = (
                (result_df['delivered_units_delta'] / result_df['delivered_units_previous'].replace({0: 1})) * 100).round(2)

        print("Данные с дельтой и процентом отклонения:", result_df.head())

        # Фильтрация по проценту изменения
        filtered_result_df = result_df[result_df['ordered_units_percent_change'].abs() > self.delta_threshold]
        print("Данные после фильтрации:", filtered_result_df.head())

        # Получение дополнительных данных асинхронно
        additional_data_url = 'http://217.25.93.96/ozon_skus_fbo?company=missyourkiss'
        additional_data = await self.get_additional_data(additional_data_url)

        additional_df = pd.DataFrame(additional_data)
        additional_df['sku'] = additional_df['sku'].astype(str)  # Преобразование SKU в строку
        additional_df = additional_df[['sku', 'offer_id', 'link', 'name', 'fbo_sku']]
        additional_df.columns = ['SKU', 'Offer ID', 'Link', 'Name', 'FBO SKU']

        # Преобразование SKU в строку
        filtered_result_df.loc[:, 'SKU'] = filtered_result_df['SKU'].astype(str)

        # Объединение дополнительных данных с основными данными
        final_df = pd.merge(filtered_result_df, additional_df, on='SKU', how='left')

        # Проверка и объединение по FBO SKU, если SKU не совпадает
        missing_data = final_df[final_df['Offer ID'].isna()]
        if not missing_data.empty:
            missing_skus = missing_data['SKU'].tolist()
            # Преобразование FBO SKU в строку перед сравнением
            additional_df['FBO SKU'] = additional_df['FBO SKU'].astype(str)
            additional_fbo_df = additional_df[additional_df['FBO SKU'].isin(missing_skus)]
            additional_fbo_df = additional_fbo_df[['FBO SKU', 'Offer ID', 'Link', 'Name']]
            additional_fbo_df.columns = ['SKU', 'Offer ID', 'Link', 'Name']
            final_df = pd.merge(final_df, additional_fbo_df, on='SKU', how='left', suffixes=('', '_fbo'))
            final_df['Offer ID'] = final_df['Offer ID'].combine_first(final_df['Offer ID_fbo'])
            final_df['Link'] = final_df['Link'].combine_first(final_df['Link_fbo'])
            final_df['Name'] = final_df['Name'].combine_first(final_df['Name_fbo'])
            final_df = final_df.drop(columns=['Offer ID_fbo', 'Link_fbo', 'Name_fbo'])

        # Переупорядочение столбцов
        final_df = final_df[['SKU', 'Offer ID', 'Link', 'Name', 'ordered_units_current', 'ordered_units_previous',
                             'ordered_units_delta', 'ordered_units_percent_change', 'delivered_units_current',
                             'delivered_units_previous', 'delivered_units_delta', 'delivered_units_percent_change']]

        # Переименование столбцов
        final_df.columns = ['nmID', 'Артикул', 'Ссылка', 'Название', f'Заказы текущий период ({current_period_str})',
                            f'Заказы предыдущий период ({previous_period_str})', 'Дельта заказов', 'Процент отклонения заказов',
                            f'Выкупы текущий период ({current_period_str})', f'Выкупы предыдущий период ({previous_period_str})',
                            'Дельта выкупов', 'Процент отклонения выкупов']

        # Сохранение итогового результата в Excel
        output_file_path = f"{current_period_str} - Ozon Delta.xlsx"
        final_df.to_excel(output_file_path, index=False)
        print(f"Файл сохранен: {output_file_path}")

        # Открытие созданного Excel файла для добавления стилей
        wb = load_workbook(output_file_path)
        ws = wb.active

        # Определение цветов для заливки
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

        # Применение заливки для процентного отклонения заказов
        for row in ws.iter_rows(min_row=2, min_col=8, max_col=8, max_row=ws.max_row):
            for cell in row:
                if cell.value > 0:
                    cell.fill = green_fill
                elif cell.value < 0:
                    cell.fill = red_fill

        # Применение заливки для процентного отклонения выкупов
        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12, max_row=ws.max_row):
            for cell in row:
                if cell.value > 0:
                    cell.fill = green_fill
                elif cell.value < 0:
                    cell.fill = red_fill

        # Сохранение файла
        wb.save(output_file_path)

        return output_file_path
