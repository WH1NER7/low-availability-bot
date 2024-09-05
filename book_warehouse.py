import time

import openpyxl
import requests
import json
import base64
from datetime import datetime

COLOR_ORDER = [
    '#FF0000',
    '#FF9900',
    '#FFFF00',
    '#00FF00',
    '#00FFFF',
    '#4A86E8',
    '#0000FF',
    '#9900FF',
    '#FF00FF',
    '#660000',
    '#7F6000',
    '#0C343D',
    '#20124D',
    '#E06666',
    '#FFD966',
    '#93C47D',
    '#8E7CC3',
    '#C27BA0',
    '#F6B26B',
    '#A2C4C9'
]


def encode_excel_to_base64(file_path):
    with open(file_path, "rb") as file:
        encoded_string = base64.b64encode(file.read()).decode('utf-8')
    return encoded_string


def decode_base64_to_excel(encoded_string, output_path):
    with open(output_path, "wb") as file:
        file.write(base64.b64decode(encoded_string))


def get_colors_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    colors = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:  # Проверка на наличие данных в ячейке
                fill = cell.fill
                if fill.start_color.rgb:
                    color = fill.start_color.rgb[-6:]  # Получаем последние 6 символов для hex-кода цвета
                    color = "#" + color  # Преобразуем в формат hex
                    if color in COLOR_ORDER:
                        colors[cell.coordinate] = color
    unique_colors = list(set(colors.values()))
    print("Extracted unique colors:", unique_colors)
    return colors, unique_colors


def add_columns_and_fill_colors(file_path, barcodes, unique_colors):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    num_columns = ws.max_column

    # Переименовываем первые две колонки
    ws.cell(row=1, column=1, value="баркод товара")
    ws.cell(row=1, column=2, value="кол-во товаров")

    # Вставляем 2 новые колонки после существующих
    ws.insert_cols(num_columns + 1)
    ws.insert_cols(num_columns + 2)
    ws.cell(row=1, column=num_columns + 1, value="шк короба")
    ws.cell(row=1, column=num_columns + 2, value="срок годности")

    # Сортируем unique_colors в соответствии с COLOR_ORDER
    color_order_clean = [color[1:] for color in COLOR_ORDER]
    color_index = {color: index for index, color in enumerate(color_order_clean)}
    sorted_colors = sorted([color[1:] for color in unique_colors if color[1:] in color_index], key=lambda color: color_index[color])

    # Соотносим цвета и баркоды
    color_to_barcode = {color: barcodes[i] for i, color in enumerate(sorted_colors)}
    print("Color to barcode mapping:", color_to_barcode)

    # Заполняем 3ю колонку в зависимости от цвета во второй колонке
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value is not None:  # Только ячейки с данными
                fill = cell.fill
                if fill and fill.start_color and fill.start_color.rgb:
                    color = fill.start_color.rgb[-6:]  # Получаем последние 6 символов для hex-кода цвета
                    color = "#" + color  # Преобразуем в формат hex
                    if color not in COLOR_ORDER:
                        print(f"Unexpected color found: {color} in cell {cell.coordinate}")
                    else:
                        barcode = color_to_barcode.get(color[1:], "")
                        ws.cell(row=cell.row, column=num_columns + 1, value=barcode)

    # Сохраняем изменения и возвращаем путь к обновленному файлу
    updated_file_path = "updated_" + file_path
    wb.save(updated_file_path)
    return updated_file_path


def log_error(error_message, status_code, params):
    log_entry = {
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": status_code,
        "error": error_message,
        "params": params
    }
    print("Error:", log_entry)
    with open("error_log.txt", "a") as log_file:
        log_file.write(json.dumps(log_entry) + "\n")


def create_draft(session, headers): # ТУТ НОРМ
    url = "https://seller-supply.wildberries.ru/ns/sm-draft/supply-manager/api/v1/draft/create"
    payload = {"params": {}, "jsonrpc": "2.0", "id": "json-rpc_24"}
    response = session.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        return response.json().get('result').get('draftID')
    else:
        log_error("Ошибка при создании поставки", response.status_code, payload)
        raise Exception(f"Failed to create draft: {response.text}")


import pandas as pd


def excel_to_list_of_dicts(file_path):
    # Читаем данные из Excel, используя имена столбцов "Баркод" и "Количество"
    df = pd.read_excel(file_path, usecols=["Баркод", "Количество"])

    # Переименовываем столбцы для нужного формата
    df.rename(columns={"Баркод": "barcode", "Количество": "quantity"}, inplace=True)

    # Преобразуем данные в список словарей
    list_of_dicts = df.to_dict('records')

    return list_of_dicts


def send_draft_update(session, headers, barcodes_list, draftID):
    url = "https://seller-supply.wildberries.ru/ns/sm-draft/supply-manager/api/v1/draft/UpdateDraftGoods"

    # Формируем тело запроса
    payload = {
        "params": {
            "draftID": draftID,
            "barcodes": barcodes_list
        },
        "jsonrpc": "2.0",
        "id": "json-rpc_24"
    }

    # Отправляем запрос
    response = session.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        return True
    else:
        return False



def upload_goods_from_xls(session, headers, draftID, encoded_xls):
    url = "https://seller-supply.wildberries.ru/ns/sm-draft/supply-manager/api/v1/draft/goodsFromXLS"
    payload = {"params": {"draftID": draftID, "xlsBytes": encoded_xls}, "jsonrpc": "2.0", "id": "json-rpc_37"}
    response = session.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        return True
    else:
        log_error("Ошибка. Не удалось загрузить эксель", response.status_code, payload)
        raise Exception(f"Failed to upload goods from XLS: {response.text}")


def create_supply(session, headers, draftID, warehouseID):
    url = "https://seller-supply.wildberries.ru/ns/sm-supply/supply-manager/api/v1/supply/create"
    payload = {"params": {"boxTypeMask": 4, "draftID": draftID, "transitWarehouseId": None, "warehouseId": warehouseID},
               "jsonrpc": "2.0", "id": "json-rpc_30"}
    response = session.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        return response.json().get('result').get('ids')[0].get('Id')
    else:
        log_error("Ошибка при создании поставки", response.status_code, payload)
        raise Exception(f"Failed to create supply: {response.text}")


def add_many_plan(session, headers, deliveryDate, preOrderId, warehouseID):
    url = "https://seller-supply.wildberries.ru/ns/sm/supply-manager/api/v1/plan/addMany"
    payload = {"params": {"preorders": [
        {"deliveryDate": deliveryDate, "preOrderId": preOrderId, "warehouseId": warehouseID,
         "supplierAssignUUID": None}]}, "jsonrpc": "2.0", "id": "json-rpc_85"}
    response = session.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        return response.json().get('result').get('out')[0].get('result').get('supplyId')
    else:
        log_error("Не удалось записать данные водителя", response.status_code, payload)
        raise Exception(f"Failed to add many plan: {response.text}")


def create_box_barcodes(session, headers, supplyId, barcodeNumber):
    url = "https://seller-supply.wildberries.ru/ns/sm-box/supply-manager/api/v1/box/createBoxBarcodes"
    payload = {"params": {"supplyId": supplyId, "barcodeNumber": barcodeNumber}, "jsonrpc": "2.0", "id": "json-rpc_224"}
    response = session.post(url, headers=headers, json=payload)
    print(response.content)
    if response.status_code == 200:
        print(response.content)
        return response.json().get('result').get('barcodes')
    else:
        log_error("Ошибка при создании баркодов", response.status_code, payload)
        raise Exception(f"Failed to create box barcodes: {response.text}")


def upload_boxes_from_xls(session, headers, incomeID, encoded_xls):
    url = "https://seller-supply.wildberries.ru/ns/sm-box/supply-manager/api/v1/box/boxesFromXLS"
    payload = {"params": {"incomeID": incomeID, "xlsBytes": encoded_xls}, "jsonrpc": "2.0", "id": "json-rpc_283"}
    response = session.post(url, headers=headers, json=payload)

    if response.status_code == 200:
        return True
    else:
        log_error("Ошибка при загрузке боксов из эксель", response.status_code, payload)
        raise Exception(f"Failed to upload boxes from XLS: {response.text}")


def get_trn_details(session, headers, supplyId):
    url = "https://seller-supply.wildberries.ru/ns/sm/supply-manager/api/v1/barcode"
    payload = [{"method": "TRNDetails", "params": {"supplyId": supplyId}, "id": "json-rpc_88", "jsonrpc": "2.0"}]
    response = session.post(url, headers=headers, json=payload)

    if response.status_code == 200:
        return response.json()[0].get('result').get('details').get('trns')[0].get('barcode').get('barcodeId')
    else:
        log_error("Ошибка при получении деталей TRN", response.status_code, payload)
        raise Exception(f"Failed to get TRN details: {response.text}")


def set_trn_details(session, headers, barcodeId, supplyId, number):
    url = "https://seller-supply.wildberries.ru/ns/sm/supply-manager/api/v1/barcode/setTRNDetails"
    payload = {"params": {"barcodeId": barcodeId, "boxTypeName": "box", "barcodePrefix": "WB-GI-", "firstName": "Тест ",
                          "lastName": "Тест", "carModel": "Фольсваген Кэдди", "carNumber": "P181PB116",
                          "supplyId": supplyId,
                          "number": number, "supplierAssignUUID": None, "phone": "79512428781"}, "jsonrpc": "2.0",
               "id": "json-rpc_50"}
    response = session.post(url, headers=headers, json=payload)
    if response.status_code == 200:
        return True
    else:
        log_error("Ошибка при установке деталей TRN", response.status_code, payload)
        raise Exception(f"Failed to set TRN details: {response.text}")


def process_supply(date, warehouseID, excel_file_path, cookie):
    # Заголовки для всех запросов
    headers = {
        'Content-Type': 'application/json',
        'Accept': 'application/json',
        'Cookie': cookie,
        "Accept-Encoding": "gzip, deflate, br",
        "User-Agent": "PostmanRuntime/7.28.4",
    }

    # Инициализация сессии
    session = requests.Session()

    # Получение цветов из Excel файла
    colors, unique_colors = get_colors_from_excel(excel_file_path)

    # Создание черновика и загрузка товаров из Excel
    draftID = create_draft(session, headers)  # api/v1/draft/create

    # новый блок возможно нужен
    barcodes_list = excel_to_list_of_dicts(excel_file_path)
    draft_status_upd = send_draft_update(session, headers, barcodes_list, draftID)
    if draft_status_upd:
        encoded_xls = encode_excel_to_base64(excel_file_path)
        upload_goods_from_xls(session, headers, draftID, encoded_xls)  # api/v1/draft/goodsFromXLS

        # Создание поставки и генерация баркодов коробок
        preorder_id = create_supply(session, headers, draftID, warehouseID)  # /api/v1/supply/create

        # addMany
        supplyId = add_many_plan(session, headers, date, preorder_id, warehouseID)

        barcodeNumber = len(unique_colors)

        barcodes = create_box_barcodes(session, headers, supplyId, barcodeNumber)  # api/v1/barcode/createBoxBarcodes

        # Обновление Excel файла с баркодами и загрузка обновленного файла
        updated_file_path = add_columns_and_fill_colors(excel_file_path, barcodes, unique_colors)

        encoded_updated_xls = encode_excel_to_base64(updated_file_path)
        upload_boxes_from_xls(session, headers, supplyId, encoded_updated_xls)  # /api/v1/box/boxesFromXLS

        # Получение деталей TRN и установка деталей TRN
        barcodeId = get_trn_details(session, headers, supplyId)  # /api/v1/barcode
        number = barcodeNumber
        supply_status = set_trn_details(session, headers, barcodeId, supplyId, number)  # /api/v1/barcode/setTRNDetails
        return supply_status, supplyId


# cookie2 = "_wbauid=9110205211698128745; ___wbu=12ae082d-b28d-44c4-b9ac-493a84ce6307.1698128748; BasketUID=dd4dccdf3ec848f1949710a632e14fea; external-locale=ru; wbx-validation-key=bead8ca0-c64d-4cf1-b194-064ea8a49f38; device_id_guru=18e08cd3ada-41b0c8de0fc5a13f; client_ip_guru=5.167.249.195; _ym_uid=1709544651663368814; _ym_d=1709544651; adult-content-wbguide=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJTaG93QWR1bHQiOmZhbHNlLCJleHAiOjE3NDEwODA2NTJ9; wb-pid=gYEi7-d4psdDu4SAR2a89L1RAAABjuvofstFpAHsjrlZFD_1YSM-9t8vaAmbNWVqUyA2s6hphGRW6A; x-supplier-id=3117fcc2-af08-5e08-835d-8a036116acd0; x-supplier-id-external=3117fcc2-af08-5e08-835d-8a036116acd0; current_feature_version=CB53E11B-0E2D-4EA0-A8AD-76385928A81D; wb-sid=b48ed8ac-3e39-4ce3-b35f-f639a28debfa; wb-id=gYGQNfBlk-5BZI_H-5vrMJKRAAABkbglm8r8wqlV4VByoKUNHH_14pkeiD4X5TlVUjQV0-PoX3RL1mI0OGVkOGFjLTNlMzktNGNlMy1iMzVmLWY2MzlhMjhkZWJmYQ; WBTokenV3=eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.eyJpYXQiOjE3MjUzNzEzODUsInZlcnNpb24iOjIsInVzZXIiOiI1MTI1MDU4MiIsInNoYXJkX2tleSI6IjEwIiwiY2xpZW50X2lkIjoic2VsbGVyLXBvcnRhbCIsInNlc3Npb25faWQiOiIwNTRkNDViOGNiM2Q0NmVkOTM1YTRhMTliMzczMTcyNyIsInVzZXJfcmVnaXN0cmF0aW9uX2R0IjoxNjc3Nzc1MzY4LCJ2YWxpZGF0aW9uX2tleSI6ImVlYzkwZjA1ZjM2NzFjNDdjYTg5YzJhZWMwMWMyODZkYjUyOTU5MGQ2ZjZlYWQ5MzQ0NWYzNTE4ZGY0NWViYjgiLCJwaG9uZSI6IiJ9.CpOi2WqKNoZTfLhtz08nThYqQakOlu01Zjo4y9oCzYaJvZGUgxR2-zPxcka01keFPib1rdn2OBfbCDmXbgOkmyMXa_oFYR9O5r0QaRe0PPXbIat1szZvGDA1kUMa1jD6_p9aaDKw-wcDAOh89ivj77DQViFez362Gbu2ZdDTvuIlkWGi1FF0dSBPwIsge-EK3iE0oXwZJBdmC2NzLzyD4CxZTpK2zudu8xHpPZ4aOKy3Us5Jt5QDOZxquOXGITtkKx7bUmlTv4WecCgHJ0rruhHRrftwLR_yURfoQ3ZwBaRbDZMz-7LhlUu9hAtLH2WRDjsQB1UDbfhu0YVQcJEw5Q; locale=ru; __zzatw-wb=MDA0dC0cTHtmcDhhDHEWTT17CT4VHThHKHIzd2UuPGciX0xhITVRP0FaW1Q4NmdBEXUmCQg3LGBwVxlRExpceEdXeiwbFnpyJFN/DVw/RWllbQwtUlFRS19/Dg4/aU5ZQ11wS3E6EmBWGB5CWgtMeFtLKRZHGzJhXkZpdRVPDA8XQUYne1xtaFFjTRZTTFhSMilKEgglJlYKDGI/SF9vG3siXyoIJGM1Xz9EaVhTMCpYQXt1J3Z+KmUzPGwfZEpgIEdVT3onHg1pN2wXPHVlLwkxLGJ5MVIvE0tsP0caRFpbQDsyVghDQE1HFF9BWncyUlFRS2EQR0lrZU5TQixmG3EVTQgNND1aciIPWzklWAgSPwsmIBN8bipQCwtdPkNzbxt/Nl0cOWMRCxl+OmNdRkc3FSR7dSYKCTU3YnAvTCB7SykWRxsyYV5GaXUVUggMYnN0KHUmbCMdZ0ReVENdSgkoHkV0c1RPChFcPkcmLl07VxlRDxZhDhYYRRcje0I3Yhk4QhgvPV8/YngiD2lIYCFIWFR6KRkTfGwqS3FPLH12X30beylOIA0lVBMhP05yOxnSJg==; __zzatw-wb=MDA0dC0cTHtmcDhhDHEWTT17CT4VHThHKHIzd2UuPGciX0xhITVRP0FaW1Q4NmdBEXUmCQg3LGBwVxlRExpceEdXeiwbFnpyJFN/DVw/RWllbQwtUlFRS19/Dg4/aU5ZQ11wS3E6EmBWGB5CWgtMeFtLKRZHGzJhXkZpdRVPDA8XQUYne1xtaFFjTRZTTFhSMilKEgglJlYKDGI/SF9vG3siXyoIJGM1Xz9EaVhTMCpYQXt1J3Z+KmUzPGwfZEpgIEdVT3onHg1pN2wXPHVlLwkxLGJ5MVIvE0tsP0caRFpbQDsyVghDQE1HFF9BWncyUlFRS2EQR0lrZU5TQixmG3EVTQgNND1aciIPWzklWAgSPwsmIBN8bipQCwtdPkNzbxt/Nl0cOWMRCxl+OmNdRkc3FSR7dSYKCTU3YnAvTCB7SykWRxsyYV5GaXUVUggMYnN0KHUmbCMdZ0ReVENdSgkoHkV0c1RPChFcPkcmLl07VxlRDxZhDhYYRRcje0I3Yhk4QhgvPV8/YngiD2lIYCFIWFR6KRkTfGwqS3FPLH12X30beylOIA0lVBMhP05yOxnSJg==; cfidsw-wb=vjz4ClLCQhdmym1MDGeZkMY7eVy4RcIaCjkeV5aAp6QNKN1jCaMV1LfjNeoRClqZAYUOoN01cXwN3/wbhesY7DOs8Zxnbb8Ty3E6JELbnOQHyAYco+vIlLKhQCNz5Hh2deKsrRHyLFrATVqyoTroYV9QvQPFYRCDsPdJbKp2Pw==; cfidsw-wb=vjz4ClLCQhdmym1MDGeZkMY7eVy4RcIaCjkeV5aAp6QNKN1jCaMV1LfjNeoRClqZAYUOoN01cXwN3/wbhesY7DOs8Zxnbb8Ty3E6JELbnOQHyAYco+vIlLKhQCNz5Hh2deKsrRHyLFrATVqyoTroYV9QvQPFYRCDsPdJbKp2Pw=="
# process_supply("2024-09-10T00:00:00Z", 1733, "test_book.xlsx", cookie2)